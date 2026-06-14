"""
file_parser_generic.py
----------------------
Best-effort parser for arbitrary tabular files.

Supports:
  .xlsx / .xls  — openpyxl required (pip install openpyxl)
  .csv          — comma-delimited
  .tsv          — tab-delimited
  .txt / .dat / other — delimiter sniffed, then numpy genfromtxt fallback

The key feature is *sub-table detection*: a single sheet or file can contain
multiple independent tables laid out side-by-side (separated by empty columns)
or stacked (separated by empty rows).  Each block is returned as a separate
GenericTable so the user can choose which one to plot.
"""

from __future__ import annotations

import csv
import os
from dataclasses import dataclass, field
from typing import Optional

import numpy as np


# ---------------------------------------------------------------------------
# Public data structure
# ---------------------------------------------------------------------------

@dataclass
class GenericTable:
    name:     str
    headers:  list[str]
    data:     np.ndarray        # (n_rows, n_cols) float64; NaN for missing
    metadata: dict = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def load_any_file(path: str) -> list[GenericTable]:
    """
    Parse *path* and return every detected table.
    Raises ValueError if nothing useful is found.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xls'):
        return _parse_excel(path)
    elif ext == '.csv':
        return _parse_delimited(path, delimiter=',')
    elif ext == '.tsv':
        return _parse_delimited(path, delimiter='\t')
    else:
        return _parse_text(path)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _parse_excel(path: str) -> list[GenericTable]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Install openpyxl to read .xlsx files:  pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    tables: list[GenericTable] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Load full sheet into a 2-D list, trim trailing all-None rows
        grid: list[list] = [list(row) for row in ws.iter_rows(values_only=True)]
        while grid and all(c is None for c in grid[-1]):
            grid.pop()
        if not grid:
            continue

        n_rows = len(grid)
        n_cols = max(len(r) for r in grid)
        for r in grid:                          # pad to uniform width
            while len(r) < n_cols:
                r.append(None)

        # ── Find empty-column boundaries ──────────────────────────────────
        empty_col = [
            all(grid[r][c] is None for r in range(n_rows))
            for c in range(n_cols)
        ]

        # Collect contiguous non-empty column spans
        col_groups: list[tuple[int, int]] = []   # (start_inclusive, end_exclusive)
        c = 0
        while c < n_cols:
            if not empty_col[c]:
                j = c
                while j < n_cols and not empty_col[j]:
                    j += 1
                col_groups.append((c, j))
                c = j
            else:
                c += 1

        for c_start, c_end in col_groups:
            sub = _extract_block(grid, n_rows, c_start, c_end)
            if sub is None:
                continue

            col_letter_range = f"cols {_col_letter(c_start)}-{_col_letter(c_end - 1)}"
            tname = (sheet_name if len(col_groups) == 1
                     else f"{sheet_name} | {col_letter_range}")
            t = _block_to_table(sub, tname, {'sheet': sheet_name, 'col_start': c_start})
            if t is not None:
                tables.append(t)

    return tables


def _extract_block(grid: list[list], n_rows: int,
                   c_start: int, c_end: int) -> Optional[list[list]]:
    """Slice columns c_start..c_end-1 from grid; trim trailing all-None rows."""
    sub = [[grid[r][c] for c in range(c_start, c_end)] for r in range(n_rows)]
    while sub and all(c is None for c in sub[-1]):
        sub.pop()
    return sub if len(sub) >= 2 else None


def _block_to_table(sub: list[list], name: str, meta: dict) -> Optional[GenericTable]:
    """
    Given a 2-D block of raw cells, find the header row, extract numeric data,
    drop all-NaN rows and columns, and return a GenericTable.
    """
    # ── Find header row ────────────────────────────────────────────────────
    # The header row is the LAST row where >= 50 % of non-None values are strings.
    header_row = 0
    for ri, row in enumerate(sub):
        non_none = [c for c in row if c is not None]
        if not non_none:
            continue
        str_frac = sum(1 for c in non_none if isinstance(c, str)) / len(non_none)
        if str_frac >= 0.5:
            header_row = ri

    # ── Build column headers ───────────────────────────────────────────────
    headers: list[str] = []
    for ci, val in enumerate(sub[header_row]):
        if val is not None and str(val).strip():
            headers.append(str(val).strip())
        else:
            # Fall back to nearest non-None value above in the same column
            label = f"Col{ci + 1}"
            for prev_ri in range(header_row - 1, -1, -1):
                pv = sub[prev_ri][ci]
                if pv is not None and str(pv).strip():
                    label = str(pv).strip()
                    break
            headers.append(label)

    # ── Convert data rows to float ─────────────────────────────────────────
    data_rows = sub[header_row + 1:]
    if not data_rows:
        return None

    float_grid: list[list[float]] = []
    for row in data_rows:
        converted = []
        for v in row:
            try:
                converted.append(float(v))   # type: ignore[arg-type]
            except (TypeError, ValueError):
                converted.append(float('nan'))
        float_grid.append(converted)

    arr = np.array(float_grid, dtype=np.float64)

    # Drop all-NaN columns, then all-NaN rows
    valid_cols = [c for c in range(arr.shape[1]) if not np.all(np.isnan(arr[:, c]))]
    if not valid_cols:
        return None
    arr = arr[:, valid_cols]
    headers = [headers[c] if c < len(headers) else f"Col{c + 1}" for c in valid_cols]

    valid_rows = [r for r in range(arr.shape[0]) if not np.all(np.isnan(arr[r, :]))]
    if not valid_rows:
        return None
    arr = arr[valid_rows, :]

    return GenericTable(name=name, headers=headers, data=arr, metadata=meta)


def _col_letter(i: int) -> str:
    """0-based column index → Excel column letter (A, B, … Z, AA, AB, …)."""
    s = ''
    i += 1
    while i > 0:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


# ---------------------------------------------------------------------------
# CSV / TSV
# ---------------------------------------------------------------------------

def _parse_delimited(path: str, delimiter: str) -> list[GenericTable]:
    with open(path, 'r', encoding='utf-8', errors='replace') as fh:
        rows = [r for r in csv.reader(fh, delimiter=delimiter)]
    if not rows:
        return []
    t = _rows_to_table(rows, os.path.basename(path))
    return [t] if t is not None else []


# ---------------------------------------------------------------------------
# Plain text (sniff delimiter)
# ---------------------------------------------------------------------------

def _parse_text(path: str) -> list[GenericTable]:
    with open(path, 'r', encoding='utf-8', errors='replace') as fh:
        sample = fh.read(8192)

    # Pick the most-frequent candidate delimiter
    counts = {'\t': sample.count('\t'), ',': sample.count(','), ';': sample.count(';')}
    best = max(counts, key=counts.get)
    if counts[best] > 0:
        return _parse_delimited(path, best)

    # Last resort: numpy whitespace splitting
    try:
        arr = np.genfromtxt(path, comments='#', filling_values=np.nan)
        if arr.ndim == 1:
            arr = arr[:, np.newaxis]
        headers = [f"Col{i + 1}" for i in range(arr.shape[1])]
        return [GenericTable(name=os.path.basename(path), headers=headers, data=arr)]
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Helpers shared by CSV / text paths
# ---------------------------------------------------------------------------

def _rows_to_table(rows: list[list[str]], name: str) -> Optional[GenericTable]:
    if not rows:
        return None

    # Find header row: last row where >= 50 % of non-empty cells are non-numeric
    header_row = 0
    for ri, row in enumerate(rows):
        non_empty = [c for c in row if c.strip()]
        if not non_empty:
            continue
        non_numeric = sum(1 for c in non_empty if not _is_numeric(c))
        if non_numeric / len(non_empty) >= 0.5:
            header_row = ri

    headers = [c.strip() or f"Col{i + 1}" for i, c in enumerate(rows[header_row])]

    float_grid: list[list[float]] = []
    for row in rows[header_row + 1:]:
        if not any(c.strip() for c in row):
            continue
        converted = []
        for c in row:
            try:
                converted.append(float(c))
            except ValueError:
                converted.append(float('nan'))
        float_grid.append(converted)

    if not float_grid:
        return None

    arr = np.array(float_grid, dtype=np.float64)
    return GenericTable(name=name, headers=headers, data=arr)


def _is_numeric(s: str) -> bool:
    try:
        float(s)
        return True
    except ValueError:
        return False
